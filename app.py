"""
Geo PDF Pipeline — Compare Claude Vision vs Google Cloud Vision OCR.
Upload an image → run both pipelines → compare results side-by-side → export to Excel.
"""

import os
import json
import time
import base64
import uuid
import tempfile
from pathlib import Path
from datetime import datetime
from io import BytesIO

from flask import Flask, request, jsonify, send_from_directory, send_file
from dotenv import load_dotenv
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()

app = Flask(__name__, static_folder="static")

UPLOAD_DIR = Path("uploads")
RESULTS_DIR = Path("results")
UPLOAD_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

# In-memory store
documents = {}

# Runtime API keys (set via UI or .env)
runtime_keys = {
    "anthropic": os.environ.get("ANTHROPIC_API_KEY", ""),
    "google": os.environ.get("GOOGLE_API_KEY", ""),
}

ALLOWED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tiff", ".tif", ".bmp", ".webp", ".pdf"}


def image_to_base64(filepath: str) -> str:
    """Read an image file and return base64-encoded PNG."""
    img = Image.open(filepath)
    if img.mode == "RGBA":
        img = img.convert("RGB")
    buf = BytesIO()
    img.save(buf, format="PNG")
    return base64.standard_b64encode(buf.getvalue()).decode("utf-8")


def pdf_to_base64_pages(filepath: str) -> list[str]:
    """Convert PDF pages to base64-encoded PNGs."""
    from pdf2image import convert_from_path
    images = convert_from_path(filepath, dpi=300)
    encoded = []
    for img in images:
        buf = BytesIO()
        img.save(buf, format="PNG")
        encoded.append(base64.standard_b64encode(buf.getvalue()).decode("utf-8"))
    return encoded


# ─── Claude Vision Pipeline ───

def extract_claude(image_b64: str, api_key: str) -> dict:
    """Extract text using Claude Vision."""
    import anthropic

    client = anthropic.Anthropic(api_key=api_key)
    start = time.time()

    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": "image/png", "data": image_b64},
                },
                {
                    "type": "text",
                    "text": """Extract ALL text from this image. Include both typed/printed text AND handwritten text.

Return a JSON object:
{
  "sections": [
    {
      "type": "header" | "field" | "table" | "paragraph" | "handwritten_note" | "label" | "other",
      "label": "<field label if applicable>",
      "content": "<the extracted text>",
      "is_handwritten": <boolean>,
      "confidence": "high" | "medium" | "low"
    }
  ],
  "tables": [
    {
      "title": "<table title if visible>",
      "headers": ["col1", "col2"],
      "rows": [["val1", "val2"]]
    }
  ],
  "raw_text": "<all text in reading order>"
}

Be thorough. Capture every piece of text — stamps, signatures, marginal notes, form labels AND filled-in values. For handwritten text, flag confidence level.""",
                },
            ],
        }],
    )

    elapsed = round(time.time() - start, 2)
    text = response.content[0].text

    # Parse JSON from response
    try:
        start_idx = text.find("{")
        end_idx = text.rfind("}") + 1
        if start_idx >= 0 and end_idx > start_idx:
            result = json.loads(text[start_idx:end_idx])
            result["processing_time"] = elapsed
            result["pipeline"] = "claude_vision"
            result["model"] = "claude-sonnet-4-20250514"
            result["input_tokens"] = response.usage.input_tokens
            result["output_tokens"] = response.usage.output_tokens
            return result
    except json.JSONDecodeError:
        pass

    return {
        "sections": [{"type": "paragraph", "content": text, "is_handwritten": False, "confidence": "medium"}],
        "tables": [],
        "raw_text": text,
        "processing_time": elapsed,
        "pipeline": "claude_vision",
        "model": "claude-sonnet-4-20250514",
    }


# ─── Google Cloud Vision Pipeline ───

def extract_google(image_b64: str, api_key: str) -> dict:
    """Extract text using Google Cloud Vision API (REST, no service account needed)."""
    import urllib.request
    import urllib.error

    start = time.time()

    payload = {
        "requests": [{
            "image": {"content": image_b64},
            "features": [
                {"type": "DOCUMENT_TEXT_DETECTION"},
                {"type": "TEXT_DETECTION"},
            ],
        }]
    }

    url = f"https://vision.googleapis.com/v1/images:annotate?key={api_key}"
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )

    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        error_body = e.read().decode("utf-8")
        return {
            "sections": [],
            "tables": [],
            "raw_text": "",
            "processing_time": round(time.time() - start, 2),
            "pipeline": "google_cloud_vision",
            "error": f"Google API error {e.code}: {error_body}",
        }

    elapsed = round(time.time() - start, 2)

    response_data = data.get("responses", [{}])[0]
    full_text_annotation = response_data.get("fullTextAnnotation", {})
    raw_text = full_text_annotation.get("text", "")

    # Parse into structured sections from pages/blocks
    sections = []
    tables_found = []

    pages = full_text_annotation.get("pages", [])
    for page in pages:
        for block in page.get("blocks", []):
            block_type = block.get("blockType", "TEXT")
            block_text_parts = []

            for paragraph in block.get("paragraphs", []):
                para_text_parts = []
                para_confidence = paragraph.get("confidence", 0)

                for word in paragraph.get("words", []):
                    word_text = "".join(
                        symbol.get("text", "") for symbol in word.get("symbols", [])
                    )
                    para_text_parts.append(word_text)

                block_text_parts.append(" ".join(para_text_parts))

            block_text = "\n".join(block_text_parts)
            if not block_text.strip():
                continue

            # Determine confidence level
            avg_confidence = block.get("confidence", 0)
            if not avg_confidence:
                # Calculate from paragraphs
                confidences = [p.get("confidence", 0) for p in block.get("paragraphs", [])]
                avg_confidence = sum(confidences) / len(confidences) if confidences else 0

            if avg_confidence >= 0.9:
                conf = "high"
            elif avg_confidence >= 0.7:
                conf = "medium"
            else:
                conf = "low"

            section_type = "table" if block_type == "TABLE" else "paragraph"

            sections.append({
                "type": section_type,
                "label": "",
                "content": block_text,
                "is_handwritten": avg_confidence < 0.85,
                "confidence": conf,
                "google_confidence": round(avg_confidence, 3),
            })

    # Also extract individual text annotations for comparison
    text_annotations = response_data.get("textAnnotations", [])

    return {
        "sections": sections,
        "tables": tables_found,
        "raw_text": raw_text,
        "processing_time": elapsed,
        "pipeline": "google_cloud_vision",
        "model": "Google Cloud Vision API",
        "total_blocks": len(sections),
        "text_annotations_count": len(text_annotations),
    }


def results_to_excel(doc_data: dict, output_path: str):
    """Export comparison results to Excel."""
    wb = Workbook()

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for pipeline_name, color in [("Claude Vision", "2B5B84"), ("Google Cloud Vision", "1A7D3F")]:
        key = "claude" if "Claude" in pipeline_name else "google"
        result = doc_data.get("results", {}).get(key, {})
        if not result or result.get("error"):
            continue

        ws = wb.active if pipeline_name == "Claude Vision" else wb.create_sheet()
        ws.title = pipeline_name
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        hw_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        headers = ["Type", "Label", "Content", "Handwritten", "Confidence"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = fill
            cell.border = thin_border

        row = 2
        for section in result.get("sections", []):
            ws.cell(row=row, column=1, value=section.get("type", "")).border = thin_border
            ws.cell(row=row, column=2, value=section.get("label", "")).border = thin_border
            c = ws.cell(row=row, column=3, value=section.get("content", ""))
            c.border = thin_border
            c.alignment = cell_align
            is_hw = section.get("is_handwritten", False)
            ws.cell(row=row, column=4, value="Yes" if is_hw else "No").border = thin_border
            ws.cell(row=row, column=5, value=section.get("confidence", "")).border = thin_border
            if is_hw:
                for cc in range(1, 6):
                    ws.cell(row=row, column=cc).fill = hw_fill
            for cc in range(1, 6):
                ws.cell(row=row, column=cc).font = cell_font
            row += 1

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12

    # Raw text comparison sheet
    ws_raw = wb.create_sheet("Raw Text Comparison")
    ws_raw.cell(row=1, column=1, value="Claude Vision Raw Text").font = Font(bold=True, size=12)
    ws_raw.cell(row=1, column=2, value="Google Cloud Vision Raw Text").font = Font(bold=True, size=12)
    ws_raw.column_dimensions["A"].width = 80
    ws_raw.column_dimensions["B"].width = 80

    claude_raw = doc_data.get("results", {}).get("claude", {}).get("raw_text", "")
    google_raw = doc_data.get("results", {}).get("google", {}).get("raw_text", "")
    ws_raw.cell(row=2, column=1, value=claude_raw).alignment = Alignment(wrap_text=True, vertical="top")
    ws_raw.cell(row=2, column=2, value=google_raw).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)


# ─── Routes ───

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/keys", methods=["GET", "POST"])
def manage_keys():
    if request.method == "POST":
        data = request.json
        if data.get("anthropic"):
            runtime_keys["anthropic"] = data["anthropic"]
        if data.get("google"):
            runtime_keys["google"] = data["google"]
        return jsonify({"ok": True})

    return jsonify({
        "anthropic": bool(runtime_keys["anthropic"] and runtime_keys["anthropic"] != "sk-ant-xxxxx"),
        "google": bool(runtime_keys["google"]),
    })


@app.route("/api/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": f"Unsupported format. Use: {', '.join(ALLOWED_EXTENSIONS)}"}), 400

    doc_id = str(uuid.uuid4())[:8]
    filename = f"{doc_id}_{file.filename}"
    filepath = UPLOAD_DIR / filename
    file.save(filepath)

    documents[doc_id] = {
        "id": doc_id,
        "filename": file.filename,
        "filepath": str(filepath),
        "is_pdf": ext == ".pdf",
        "status": "uploaded",
        "uploaded_at": datetime.now().isoformat(),
        "results": {},
    }

    return jsonify({"id": doc_id, "filename": file.filename, "status": "uploaded"})


@app.route("/api/process/<doc_id>", methods=["POST"])
def process(doc_id):
    if doc_id not in documents:
        return jsonify({"error": "Document not found"}), 404

    doc = documents[doc_id]
    if doc["status"] == "processing":
        return jsonify({"error": "Already processing"}), 409

    doc["status"] = "processing"
    results = {}

    # Get image as base64
    try:
        if doc["is_pdf"]:
            pages = pdf_to_base64_pages(doc["filepath"])
            image_b64 = pages[0]  # First page for now
        else:
            image_b64 = image_to_base64(doc["filepath"])
    except Exception as e:
        doc["status"] = "error"
        return jsonify({"error": f"Failed to read image: {e}"}), 500

    # Run Claude Vision
    if runtime_keys["anthropic"] and runtime_keys["anthropic"] != "sk-ant-xxxxx":
        try:
            results["claude"] = extract_claude(image_b64, runtime_keys["anthropic"])
        except Exception as e:
            results["claude"] = {"error": str(e), "pipeline": "claude_vision", "sections": [], "tables": [], "raw_text": ""}
    else:
        results["claude"] = {"error": "No API key configured", "pipeline": "claude_vision", "sections": [], "tables": [], "raw_text": ""}

    # Run Google Cloud Vision
    if runtime_keys["google"]:
        try:
            results["google"] = extract_google(image_b64, runtime_keys["google"])
        except Exception as e:
            results["google"] = {"error": str(e), "pipeline": "google_cloud_vision", "sections": [], "tables": [], "raw_text": ""}
    else:
        results["google"] = {"error": "No API key configured", "pipeline": "google_cloud_vision", "sections": [], "tables": [], "raw_text": ""}

    doc["results"] = results
    doc["status"] = "complete"

    # Save JSON
    with open(RESULTS_DIR / f"{doc_id}.json", "w") as f:
        json.dump(doc, f, indent=2, default=str)

    return jsonify({"id": doc_id, "status": "complete", "results": results})


@app.route("/api/export/<doc_id>")
def export(doc_id):
    if doc_id not in documents:
        return jsonify({"error": "Document not found"}), 404

    doc = documents[doc_id]
    if doc["status"] != "complete":
        return jsonify({"error": "Not yet processed"}), 400

    excel_path = RESULTS_DIR / f"{doc_id}.xlsx"
    results_to_excel(doc, str(excel_path))

    return send_file(
        excel_path,
        as_attachment=True,
        download_name=f"{Path(doc['filename']).stem}_comparison.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/image/<doc_id>")
def serve_image(doc_id):
    if doc_id not in documents:
        return jsonify({"error": "Not found"}), 404
    return send_file(documents[doc_id]["filepath"])


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 3014))
    print(f"\n  📄 Geo PDF Pipeline running at http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=True)
